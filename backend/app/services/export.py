"""Export service for quiz results."""

import io
from typing import List, Dict

from sqlalchemy.orm import Session

from app.models.quiz_session import QuizSession
from app.models.participant import Participant
from app.models.response import Response
from app.models.question import Question


def export_results_csv(db: Session, session_id: str) -> str:
    """Export quiz results as CSV string."""
    import csv

    session = db.query(QuizSession).filter(QuizSession.id == session_id).first()
    if not session:
        return ""

    participants = (
        db.query(Participant)
        .filter(Participant.session_id == session_id)
        .order_by(Participant.rank.asc().nullslast())
        .all()
    )

    questions = (
        db.query(Question)
        .filter(Question.session_id == session_id)
        .order_by(Question.order_index)
        .all()
    )

    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    header = ["Rank", "Team Name", "Total Score", "Correct Answers", "Avg Response Time (s)"]
    for i, q in enumerate(questions):
        header.extend([f"Q{i+1} Answer", f"Q{i+1} Correct", f"Q{i+1} Time (s)", f"Q{i+1} Points"])
    writer.writerow(header)

    # Data rows
    for participant in participants:
        avg_time = (
            round(participant.total_response_time / participant.correct_answers, 2)
            if participant.correct_answers > 0
            else 0
        )

        row = [
            participant.rank or "-",
            participant.team_name,
            participant.total_score,
            participant.correct_answers,
            avg_time,
        ]

        for question in questions:
            response = (
                db.query(Response)
                .filter(
                    Response.participant_id == participant.id,
                    Response.question_id == question.id,
                )
                .first()
            )

            if response:
                row.extend([
                    response.selected_answer,
                    "Yes" if response.is_correct else "No",
                    round(response.response_time, 2),
                    response.points_awarded,
                ])
            else:
                row.extend(["No answer", "No", "-", 0])

        writer.writerow(row)

    return output.getvalue()


def export_results_excel(db: Session, session_id: str) -> bytes:
    """Export quiz results as Excel bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    session = db.query(QuizSession).filter(QuizSession.id == session_id).first()
    if not session:
        return b""

    wb = openpyxl.Workbook()

    # Leaderboard sheet
    ws = wb.active
    ws.title = "Leaderboard"

    headers = ["Rank", "Team Name", "Total Score", "Correct Answers",
               "Avg Response Time (s)", "Fastest Response (s)"]
    ws.append(headers)

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    participants = (
        db.query(Participant)
        .filter(Participant.session_id == session_id)
        .order_by(Participant.rank.asc().nullslast())
        .all()
    )

    for p in participants:
        avg_time = (
            round(p.total_response_time / p.correct_answers, 2)
            if p.correct_answers > 0
            else 0
        )
        ws.append([
            p.rank or "-",
            p.team_name,
            p.total_score,
            p.correct_answers,
            avg_time,
            round(p.fastest_response_time, 2) if p.fastest_response_time else "-",
        ])

    # Detailed responses sheet
    ws2 = wb.create_sheet("Detailed Responses")
    questions = (
        db.query(Question)
        .filter(Question.session_id == session_id)
        .order_by(Question.order_index)
        .all()
    )

    detail_headers = ["Team Name", "Question #", "Question Text", "Selected Answer",
                      "Correct Answer", "Is Correct", "Response Time (s)", "Points"]
    ws2.append(detail_headers)

    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for p in participants:
        for i, q in enumerate(questions):
            response = (
                db.query(Response)
                .filter(
                    Response.participant_id == p.id,
                    Response.question_id == q.id,
                )
                .first()
            )

            ws2.append([
                p.team_name,
                i + 1,
                q.question_text[:100],
                response.selected_answer if response else "No answer",
                q.correct_answer,
                "Yes" if (response and response.is_correct) else "No",
                round(response.response_time, 2) if response else "-",
                response.points_awarded if response else 0,
            ])

    # Auto-adjust column widths
    for ws_sheet in [ws, ws2]:
        for column in ws_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_sheet.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
